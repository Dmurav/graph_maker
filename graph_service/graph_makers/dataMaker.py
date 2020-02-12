import openpyxl


def graph_data_from_xlsx(source_file):

    dates = []
    graph_data_1 = []
    graph_data_2 = []
    graph_data_3 = []
    # Создаём в памяти объект из исходного файла
    wb = openpyxl.load_workbook(source_file)
    # Выбираем активный лист в книге
    sheet = wb.active
    # Определяем число строк и столбцов в таблице
    rows = sheet.max_row
    cols = sheet.max_column

    if cols == 4:
        # Заполняем списки для графика
        for i in range(1, rows + 1):
            dates.append(sheet.cell(row=i, column=1).value)
            graph_data_1.append(sheet.cell(row=i, column=2).value)
            graph_data_2.append(sheet.cell(row=i, column=3).value)
            graph_data_3.append(sheet.cell(row=i, column=4).value)
        return [dates, graph_data_1, graph_data_2, graph_data_3]

    elif cols == 3:
        # Заполняем списки для графика
        for i in range(1, rows + 1):
            dates.append(sheet.cell(row=i, column=1).value)
            graph_data_1.append(sheet.cell(row=i, column=2).value)
            graph_data_2.append(sheet.cell(row=i, column=3).value)
        return [dates, graph_data_1, graph_data_2]

    elif cols == 2:
        # Заполняем списки для графика
        for i in range(1, rows + 1):
            dates.append(sheet.cell(row=i, column=1).value)
            graph_data_1.append(sheet.cell(row=i, column=2).value)
        return [dates, graph_data_1]