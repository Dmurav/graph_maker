import openpyxl
import matplotlib.pyplot as plt


class GraphMemoryBuilder:
    source_file = ""
    dates = []
    graph_data_1 = []
    graph_data_2 = []

    def __init__(self, source_file):
        self.source_file = source_file

    def prepare_data(self):
        # Создаём объект с данными из Excel
        wb = openpyxl.load_workbook(self.source_file)
        # Выбираем активный лист в книге
        sheet = wb.active
        # Вычисляем число строк и колонок
        rows = sheet.max_row
        cols = sheet.max_column
        # Заполняем списки для графика
        dates = []
        graph_data_1 = []
        graph_data_2 = []

        for i in range(1, rows + 1):
            dates.append(sheet.cell(row=i, column=1).value)
            graph_data_1.append(sheet.cell(row=i, column=2).value)
            graph_data_2.append(sheet.cell(row=i, column=3).value)

        self.dates = dates
        self.graph_data_1 = graph_data_1
        self.graph_data_2 = graph_data_2


    def make_graph_memory(self):
        # Устанавливаем размер графика
        fig, ax = plt.subplots(figsize=(8, 4))

        # Устанавливаем подписи
        ax.set_title("Утилизация памяти", fontsize=16)
        ax.set_ylabel('Утилизация памяти, %', fontsize=10)
        ax.set_xlabel('Продолжительность теста, ч:мм', fontsize=10)

        fig.tight_layout()

        # Строим график
        plt.plot(self.dates, self.graph_data_1, label='Утилизация памяти')
        plt.plot(self.dates, self.graph_data_2, label='Утилизация подкачки')
        # Устанавливаем сетку
        plt.grid()
        # Устанавливаем легенду
        plt.legend()

        plt.gcf()
        # создать файл с графиком
        plt.savefig("../result/graph_memory.png")

    def show_memory_graph(self):
        plt.show()