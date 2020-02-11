import openpyxl
import matplotlib.pyplot as plt


class GraphCPUBuilder:
    source_file = ""
    dates = []
    graph_data_1 = []
    graph_data_2 = []

    def __init__(self, source_file):
        self.source_file = source_file

    def prepare_data(self):
        # Создаём объект с данными из Excel
        wb = openpyxl.load_workbook(self.source_file)
        # Выбираем активный лист в книге
        sheet = wb.active
        # Вычисляем число строк и колонок
        rows = sheet.max_row
        cols = sheet.max_column
        # Заполняем списки для графика
        dates = []
        graph_data_1 = []
        graph_data_2 = []

        for i in range(1, rows + 1):
            dates.append(sheet.cell(row=i, column=1).value)
            graph_data_1.append(sheet.cell(row=i, column=2).value)
            graph_data_2.append(sheet.cell(row=i, column=3).value)

        self.dates = dates
        self.graph_data_1 = graph_data_1
        self.graph_data_2 = graph_data_2


    def make_graph_cpu(self):
        # Устанавливаем размер графика
        fig, ax = plt.subplots(figsize=(8, 4))
        # Устанавливаем подписи
        ax.set_title("Утилизация CPU", fontsize=16)
        ax.set_ylabel('Значение коэфициента', fontsize=10)
        ax.set_xlabel('Продолжительность теста, ч:мм', fontsize=10)
        fig.tight_layout()
        # Строим график
        plt.plot(self.dates, self.graph_data_1, label='Утилизация CPU')
        plt.plot(self.dates, self.graph_data_2, label='Длина очереди CPU')
        # Устанавливаем сетку
        plt.grid()
        # Устанавливаем легенду
        plt.legend(loc='upper right')
        plt.savefig("../result/graph_cpu.png")


    def show_cpu_graph(self):
        plt.show()